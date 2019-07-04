# -*- encoding=utf8 -*-
import time
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side,Alignment
from copy import copy
print('''
测试报告的准备工作：
1.将问题列表中仅保留一行
2.问题列表下方距离表格黑色粗线仅一行

bug列表的准备工作：
1.从禅道导出bug列表，导出时修改编码格式为GBK
2.导出后将该csv文件另存为xlsx，后续导入时拖入该xlsx文件即可
''')
text1 = input('请拖入测试报告：')
text1 = text1.replace('"','')
r = load_workbook(text1)
r_sheetnames = r.sheetnames
rt = r[r_sheetnames[0]]
report = r.active
text2 = input('请拖入bug列表:')
text2 = text2.replace('"','')
b = load_workbook(text2)
b_sheetnames = b.sheetnames
bt = b[b_sheetnames[0]]
bug = b.active
hang_report = 0
lie_report = ['B','C','D','E','F','G','H','I','J','K','L']
lie = ['C','D','H','I']
lie_bug = ['A','G','I','Z']

for i in range(1,rt.max_row):
   if '问题列表' in str(report.cell(i,3).value):
      hang_report = i + 2

for i in range(1,bt.max_row - 1):
   report.insert_rows(hang_report + 1)
   m1 = 'D' + str(hang_report + i) + ':G' + str(hang_report + i)
   m2 = 'I' + str(hang_report + i) + ':K' + str(hang_report + i)
   report.merge_cells(m1)
   report.merge_cells(m2)

   for j in range(0,len(lie_report)):
      c1 = lie_report[j] + str(hang_report)
      c2 = lie_report[j] + str(hang_report + 1)
      a = report[c1].border
      report[c2].border = copy(a)
      b = report[c1].font
      report[c2].font = copy(b)
      c = report[c1].alignment
      report[c2].alignment = copy(c)

for i in range(2,bt.max_row + 1):
   for j in range(0,len(lie_bug)):
      a = lie[j] + str(hang_report + i -2)
      b = lie_bug[j] + str(i)
      bug_value = bug[b].value
      if j == 2:
         if bug_value == 1:
            bug_value = str(bug_value) + '-A'
         elif bug_value == 2:
            bug_value = str(bug_value) + '-B'
         elif bug_value == 3:
            bug_value = str(bug_value) + '-C'
         else:
            bug_value = str(bug_value) + '-D'
      report[a].value = bug_value

filename = text1.replace('.xlsx','(已导入bug列表).xlsx')
r.save(filename)
print('导入bug列表成功！')
time.sleep(3)